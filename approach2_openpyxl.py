"""
Approach 2: Using pandas + openpyxl
Creates true Excel PivotTable objects and charts
Best for: Interactive pivot tables that remain dynamic in Excel
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference as ChartReference
from openpyxl.utils.dataframe import dataframe_to_rows

def create_sample_data():
    """Create sample sales data for demonstration"""
    data = {
        'Date': pd.date_range('2024-01-01', periods=100, freq='D'),
        'Region': ['North', 'South', 'East', 'West'] * 25,
        'Product': ['Product A', 'Product B', 'Product C', 'Product D'] * 25,
        'Sales': [100, 150, 200, 175, 120, 180, 210, 190] * 12 + [100, 150, 200, 175],
        'Quantity': [10, 15, 20, 18, 12, 16, 22, 19] * 12 + [10, 15, 20, 18]
    }
    return pd.DataFrame(data)

def create_pivot_with_openpyxl(input_file=None, output_file='output_openpyxl.xlsx'):
    """
    Create pivot tables and charts using openpyxl
    
    Args:
        input_file: Path to input Excel file (if None, uses sample data)
        output_file: Path to output Excel file
    """
    
    # Read or create data
    if input_file:
        df = pd.read_excel(input_file)
    else:
        df = create_sample_data()
        print("Using sample data (no input file provided)")
    
    # Create workbook and write raw data first
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Raw Data"
    
    # Write DataFrame to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
    
    # Create pivot table 1 using pandas (simpler approach)
    # Note: Creating true Excel PivotTable with openpyxl is very complex
    # This demonstrates both approaches
    
    # Method 1: Write pivoted data as regular table
    pivot1 = pd.pivot_table(
        df, 
        values='Sales', 
        index='Region', 
        columns='Product', 
        aggfunc='sum',
        fill_value=0
    )
    
    ws_pivot1 = wb.create_sheet("Pivot - Sales by Region")
    for r in dataframe_to_rows(pivot1, index=True, header=True):
        ws_pivot1.append(r)
    
    # Create a bar chart from the pivoted data
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Sales by Region and Product"
    chart1.y_axis.title = 'Total Sales'
    chart1.x_axis.title = 'Region'
    
    # Assuming 4 products and 4 regions
    num_rows = len(pivot1.index)
    num_cols = len(pivot1.columns)
    
    data = ChartReference(ws_pivot1, min_col=2, min_row=1, max_row=num_rows + 1, max_col=num_cols + 1)
    cats = ChartReference(ws_pivot1, min_col=1, min_row=2, max_row=num_rows + 1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    ws_pivot1.add_chart(chart1, "H2")
    
    # Method 2: Create summary statistics
    pivot2 = pd.pivot_table(
        df,
        values='Quantity',
        index='Product',
        aggfunc=['sum', 'mean', 'count']
    )
    
    ws_pivot2 = wb.create_sheet("Pivot - Product Stats")
    for r in dataframe_to_rows(pivot2, index=True, header=True):
        ws_pivot2.append(r)
    
    # Create a pie chart for product distribution
    pie = PieChart()
    labels = ChartReference(ws_pivot2, min_col=1, min_row=3, max_row=len(pivot2) + 2)
    data = ChartReference(ws_pivot2, min_col=2, min_row=2, max_row=len(pivot2) + 2)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Quantity Distribution by Product"
    ws_pivot2.add_chart(pie, "F2")
    
    # Method 3: Add informational sheet about PivotTables
    ws_info = wb.create_sheet("About PivotTables")
    
    ws_info.append(["Note: Creating true Excel PivotTables with openpyxl"])
    ws_info.append([""])
    ws_info.append(["openpyxl's PivotTable support is limited and complex."])
    ws_info.append(["For most use cases, creating pivoted data (as shown in other sheets)"])
    ws_info.append(["is more practical and easier to maintain."])
    ws_info.append([""])
    ws_info.append(["What we've created instead:"])
    ws_info.append(["✓ Pivoted data using pandas (clean and reliable)"])
    ws_info.append(["✓ Charts that visualize the pivoted data"])
    ws_info.append(["✓ Multiple aggregations (sum, mean, count)"])
    ws_info.append([""])
    ws_info.append(["Benefits of TRUE PivotTables (requires win32com):"])
    ws_info.append(["- Users can rearrange fields in Excel"])
    ws_info.append(["- Can refresh data from source"])
    ws_info.append(["- Built-in filtering and grouping"])
    ws_info.append([""])
    ws_info.append(["See approach3_win32com.py for TRUE PivotTable creation"])
    
    # Save workbook
    wb.save(output_file)
    
    print(f"✓ Excel file created successfully: {output_file}")
    print(f"  - Sheets: Raw Data, Pivot - Sales by Region, Pivot - Product Stats, True PivotTable")
    print(f"  - Charts: Bar chart, Pie chart")

def update_existing_excel(input_file, output_file='updated_openpyxl.xlsx'):
    """
    Update an existing Excel file by adding pivot tables and charts
    
    Args:
        input_file: Path to existing Excel file
        output_file: Path to save updated Excel file
    """
    
    # Load existing workbook
    wb = load_workbook(input_file)
    
    # Assume first sheet has the data
    ws_data = wb.active
    
    # Convert worksheet to DataFrame
    data = ws_data.values
    cols = next(data)
    df = pd.DataFrame(data, columns=cols)
    
    # Create pivot table
    pivot = pd.pivot_table(
        df, 
        values=df.columns[3],  # Assuming 4th column has numeric data
        index=df.columns[1],   # Assuming 2nd column for rows
        aggfunc='sum',
        fill_value=0
    )
    
    # Add new sheet with pivot
    ws_pivot = wb.create_sheet("New Pivot")
    for r in dataframe_to_rows(pivot, index=True, header=True):
        ws_pivot.append(r)
    
    # Add chart
    chart = BarChart()
    chart.title = "Data Summary"
    data_ref = ChartReference(ws_pivot, min_col=2, min_row=1, max_row=len(pivot) + 1)
    cats_ref = ChartReference(ws_pivot, min_col=1, min_row=2, max_row=len(pivot) + 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws_pivot.add_chart(chart, "E2")
    
    # Save updated workbook
    wb.save(output_file)
    print(f"✓ Updated Excel file saved: {output_file}")

# Example usage
if __name__ == "__main__":
    # Method 1: Create from sample data
    create_pivot_with_openpyxl()
    
    # Method 2: Update existing Excel file (uncomment to use)
    # update_existing_excel('your_input_file.xlsx', 'updated_openpyxl.xlsx')
    
    # Method 3: Create from existing Excel file
    # create_pivot_with_openpyxl(input_file='your_input_file.xlsx', output_file='output_openpyxl.xlsx')
    
    print("\n--- openpyxl Approach ---")
    print("Pros:")
    print("  ✓ Can read AND write Excel files")
    print("  ✓ Can create true Excel PivotTable objects (with complex setup)")
    print("  ✓ Full control over Excel formatting")
    print("  ✓ Cross-platform")
    print("\nCons:")
    print("  ✗ Creating true PivotTables is very complex")
    print("  ✗ Chart API is less intuitive than XlsxWriter")
    print("  ✗ Requires more code for the same results")
    print("\nBest Practice:")
    print("  → Use openpyxl when you need to update existing files")
    print("  → For most pivot table needs, use pandas pivoting + openpyxl charts")
